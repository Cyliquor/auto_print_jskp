import json
import time
import openpyxl
import os
import requests
import math
from configparser import ConfigParser
import xlwings as xw
import traceback
from time import sleep, strftime, localtime
import datetime


CONFIGFILE = '.\\config\\config.ini'
config = ConfigParser()
config.read(CONFIGFILE)

tuo = config['base']['tuo']
send_host = config['base']['send_host']
name = config['base']['name']
url = config['base']['get_url']

"""  
len_str 字符串长度
lens   单元格能容纳长度
ws 工作表
rows 行号
"""
def s_f(len_str, lens, ws, rows):
    row = math.ceil(len_str/lens)
    if ws.row_dimensions[rows].height >= 15*row:
        return
    else:
        ws.row_dimensions[rows].height = 15*row


def print_file_(filename):
    try:
        app = xw.App(visible=False, add_book=False)  # 启动Excel程序
        workbook = app.books.open(filename)  # 打开要打印的工作簿
        workbook.api.PrintOut(Copies=1, Collate=True)  # 打印工作簿
        workbook.close()
        app.quit()
        print('开始打印...')
        print('打印成功')
    except Exception as E:
        # with open('./test.txt', "a") as f:
        #     traceback.print_exc(file=f)
        print(E)
        print('打印出错！')


def main():
    try:
        res = requests.get(url=f'{url}{name}').text
        # res = requests.get(url='http://flask3.chinagearbox.com.cn/getjskp/RPA001').text
        data = json.loads(res)
        print(data)
    except Exception as e:
        # print(e)
        time.sleep(3)
        return
    # print(type(data))
    # print(data)
    if type(data) == dict:
        lst = []
        lst.append(data)
        data = lst
    if data != None:
        wb = openpyxl.load_workbook('./jskp.xlsx')
        ws = wb.active
        i = 0
        sheetnum = 0
        VB_names = ''
        for x in range(len(data)):
            sheetnum += len(data[x]['ZDATA'])
            VB_names = VB_names + '-' + data[x]['VBELN']
        # print(sheetnum)
        for y in range(sheetnum - 1):
            newwb = wb.copy_worksheet(ws)
            newwb.title = 'Sheet' + str(y + 2)

        m = 1
        for k in range(len(data)):
            allrow = data[k]['ZDATA']
            for tmprow in allrow:
                ws = wb[f'Sheet{m}']
                ws['B2'].value = str(data[k]['VBELN'])
                ws['B3'].value = data[k]['KUNRG'].lstrip("0")
                ws['C3'].value = data[k]['NAME1']
                ws['G36'].value = data[k]['NAME_ORG1']
                s = data[k]['FKDAT']
                ws['J3'].value = s[0:4] + '.' + s[4:6] + '.' + s[6:]
                ws['C4'].value = float(tmprow['NETWR']) + float(tmprow['MWSBP'])
                ws['I4'].value = float(tmprow['NETWR'])
                ws['A6'].value = tmprow['POSNR'].lstrip("0")
                ws['B6'].value = tmprow['MATNR'].lstrip("0")
                s_f(len(tmprow['ARKTX']), 8, ws, 6)
                ws['C6'].value = tmprow['ARKTX']
                ws['E6'].value = tmprow['PART_NO']
                ws['G6'].value = float(tmprow['FKIMG'])
                ws['H6'].value = tmprow['VRKME']
                ws['I6'].value = float(tmprow['NETWR']) / float(tmprow['FKIMG'])
                ws['J6'].value = tmprow['WAERK']
                ws['K6'].value = float(tmprow['MWSBP']) / float(tmprow['FKIMG'])
                ws['C7'].value = tmprow['ZPZXX']

                j = 0
                for tmpsubrow in tmprow['ZXP']:
                    if tmpsubrow['ZJE'][-1] == '-':
                        tmp = '-' + tmpsubrow['ZJE'].lstrip(" ").replace('-', '')
                    else:
                        tmp = tmpsubrow['ZJE']

                    if float(tmpsubrow['ZCID']) != 0:
                        s_f(len(tmpsubrow['P_MODEL1']), 8, ws, 10)
                        ws['B10'].value = tmpsubrow['P_MODEL1']
                        ws['D10'].value = tmpsubrow['GEAR_RATIO']
                        ws['I10'].value = float(tmp)
                        ws['J10'].value = float(tmp)
                    else:

                        if float(tmpsubrow['CJID']) != 0:
                            s_f(len(tmpsubrow['P_NODES']), 8, ws, 11)
                            ws['B11'].value = tmpsubrow['P_NODES']
                            ws['D11'].value = tmpsubrow['GEAR_RATIO']
                            ws['I11'].value = float(tmp)
                            ws['J11'].value = float(tmp)
                        else:
                            s_f(len(tmpsubrow['P_NO']), 8, ws, j + 12)
                            ws[f'B{j + 12}'].value = tmpsubrow['P_NO']
                            ws[f'D{j + 12}'].value = tmpsubrow['GEAR_RATIO']
                            ws[f'H{j + 12}'].value = float(tmpsubrow['ZTPSL'])
                            ws[f'I{j + 12}'].value = float(tmp)
                            ws[f'J{j + 12}'].value = float(tmp) * float(tmpsubrow['ZTPSL'])
                            j += 1
                m += 1
            i += 1
        # time_end = strftime('%Y/%m/%d %H:%M', localtime())
        wb.save(f'./开票申请单/jskp开票申请单{VB_names}.xlsx')
        os.startfile(f'.\\开票申请单\\jskp开票申请单{VB_names}.xlsx')
    print(f'开票申请单输出成功（文件名：jskp开票申请单{VB_names}）')
    # print_file_(f'.\\开票申请单\\jskp开票申请单{VB_names}.xlsx')


if __name__ == "__main__":
    # start_time = datetime.datetime.now()
    # time_limit = datetime.timedelta(hours=4)
    while True:
        try:
            main()
            # current_time = datetime.datetime.now()
            # if current_time - start_time > time_limit:
            #     break
            # print(current_time)
            time.sleep(1)
        except Exception as E:
            print(E)
        # time.sleep(1)